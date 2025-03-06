import streamlit as st
import pandas as pd
import openpyxl
import graphviz
import re
import io
from pathlib import Path

# App title
st.title("📂 Multi-File Excel Dependency Analyzer")

# File uploader for multiple Excel files
uploaded_files = st.file_uploader(
    "Upload multiple Excel files", 
    type=["xlsx", "xls"], 
    accept_multiple_files=True
)

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} files uploaded successfully!")

    # Store sheet dependencies
    file_dependencies = {}

    # Read and process each uploaded file
    excel_data = {}
    file_names = [uploaded_file.name for uploaded_file in uploaded_files]  # Store full file names
    file_stems = {str(index + 1): name for index, name in enumerate(file_names)}  # Map numeric references [1], [2], etc.

    st.write("📂 Uploaded files detected:", file_names)  # Debugging: List uploaded files
    st.write("📊 Numeric File Mapping (for Excel references):", file_stems)  # Debugging: Show Excel's numeric mapping

    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name
        file_dependencies[file_name] = set()  # Ensure every file is in the dictionary

        # Read the Excel file
        file_stream = io.BytesIO(uploaded_file.read())  # Convert to BytesIO for openpyxl
        wb = openpyxl.load_workbook(file_stream, data_only=False)
        excel_data[file_name] = wb

    # Analyze formulas and detect dependencies
    for file_name, wb in excel_data.items():
        st.write(f"🔍 Scanning file: {file_name}")  # Debugging: Show file being scanned

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_text = cell.value

                        # Debugging: Show detected formula
                        st.write(f"📊 Formula found in {file_name} - {sheet}: `{formula_text}`")

                        # Extract the referenced file (whether full path, [1], or [Simple Model - Part 1.xlsx])
                        match = re.search(r"\[(.*?)\]", formula_text)
                        if match:
                            referenced_file = match.group(1)  # Extracted reference (could be filename or [1])
                            referenced_stem = Path(referenced_file).stem

                            # Debugging: Show extracted file reference
                            st.write(f"🔗 Formula references: `{referenced_file}`")

                            # Resolve numeric references like [1] to real filenames
                            if referenced_file.isdigit() and referenced_file in file_stems:
                                resolved_filename = file_stems[referenced_file]  # Map [1] -> "Simple Model - Part 1.xlsx"
                                st.write(f"🔄 Resolved `[1]` reference to `{resolved_filename}`")
                            else:
                                resolved_filename = referenced_file  # Keep original if it's a real filename

                            # Ensure the referenced file exists in uploaded files
                            for uploaded_stem, uploaded_name in file_stems.items():
                                if uploaded_stem.lower() == Path(resolved_filename).stem.lower() and uploaded_name != file_name:
                                    file_dependencies[file_name].add(uploaded_name)  # Store dependency
                                    st.write(f"✅ Link created: `{file_name}` → `{uploaded_name}`")

    # **Debugging: Show final dependencies before plotting**
    st.write("📋 Final Detected Dependencies:", file_dependencies)

    # Ensure all files appear in the flowchart (even if they have no links)
    all_files = set(file_dependencies.keys()).union(*file_dependencies.values())

    # Generate dependency flowchart
    st.write("### 🔄 Spreadsheet Dependency Flowchart")
    flow = graphviz.Digraph()

    # **Ensure all files appear in Graphviz, even isolated ones**
    for file in all_files:
        flow.node(file)

    # **Force Graphviz to draw arrows**
    has_edges = False  # Track if at least one arrow is drawn
    for file, dependencies in file_dependencies.items():
        for dependency in dependencies:
            flow.edge(dependency, file)  # Draw arrows
            has_edges = True  # Track that we have edges

    # **If no edges were added, force a dummy edge to prevent an empty graph**
    if not has_edges:
        flow.node("No Dependencies Found", shape="plaintext")

    # Display the flowchart
    st.graphviz_chart(flow)

    # Show detected dependencies
    st.write("### 📊 Dependency Table")
    dependency_df = pd.DataFrame(
        [(file, dep) for file, deps in file_dependencies.items() for dep in deps],
        columns=["File", "Depends On"]
    )

    if dependency_df.empty:
        st.write("✅ No direct dependencies found between uploaded Excel files.")
    else:
        st.dataframe(dependency_df)
